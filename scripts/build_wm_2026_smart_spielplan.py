#!/usr/bin/env python3
"""Build an editorial smart workbook for the FIFA World Cup 2026."""

from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import datetime, timedelta, time
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


PROJECT_DIR = Path(__file__).resolve().parent
MATCHES_PATH = PROJECT_DIR / "wm_2026_matches_fifa.json"
RANKING_PATH = PROJECT_DIR / "fifa_mens_ranking_latest.json"
OUTPUT_PATH = PROJECT_DIR / "WM_2026_Smart_Spielplan.xlsx"
FLAG_DIR = PROJECT_DIR / "flags"
THUMB_DIR = FLAG_DIR / "_thumbs"
TIMEZONE = "Europe/Berlin"
MATCH_BLOCK_MINUTES = 105

UEFA_CODES = {
    "ALB", "AND", "ARM", "AUT", "AZE", "BEL", "BIH", "BLR", "BUL", "CRO",
    "CYP", "CZE", "DEN", "ENG", "ESP", "EST", "FRO", "FIN", "FRA", "GEO",
    "GER", "GIB", "GRE", "HUN", "IRL", "ISL", "ISR", "ITA", "KAZ", "KOS",
    "LIE", "LTU", "LUX", "LVA", "MDA", "MKD", "MLT", "MNE", "NED", "NIR",
    "NOR", "POL", "POR", "ROU", "RUS", "SCO", "SMR", "SRB", "SUI", "SVK",
    "SVN", "SWE", "TUR", "UKR", "WAL",
}

GERMAN_DAYS = {
    "Monday": "Montag",
    "Tuesday": "Dienstag",
    "Wednesday": "Mittwoch",
    "Thursday": "Donnerstag",
    "Friday": "Freitag",
    "Saturday": "Samstag",
    "Sunday": "Sonntag",
}


def localized_text(items: list[dict] | None, default: str = "") -> str:
    for item in items or []:
        if str(item.get("Locale", "")).lower().startswith("en"):
            return item.get("Description", default)
    return items[0].get("Description", default) if items else default


def team_name(match: dict, side: str) -> str:
    team = match.get(side)
    if team:
        return localized_text(team.get("TeamName"), team.get("ShortClubName", ""))
    return match.get("PlaceHolderA" if side == "Home" else "PlaceHolderB") or ""


def team_code(match: dict, side: str) -> str:
    team = match.get(side)
    return team.get("IdCountry", "") if team else ""


def stage_name(match: dict) -> str:
    return localized_text(match.get("StageName"))


def group_name(match: dict) -> str:
    return localized_text(match.get("GroupName"))


def stadium_name(match: dict) -> str:
    return localized_text((match.get("Stadium") or {}).get("Name"))


def city_name(match: dict) -> str:
    return localized_text((match.get("Stadium") or {}).get("CityName"))


def make_thumbnail(src: Path, width: int = 34, height: int = 24) -> Path:
    THUMB_DIR.mkdir(parents=True, exist_ok=True)
    thumb = THUMB_DIR / src.name
    if thumb.exists() and thumb.stat().st_size > 0:
        return thumb
    with PILImage.open(src) as img:
        img = img.convert("RGBA")
        img.thumbnail((width, height), PILImage.LANCZOS)
        canvas = PILImage.new("RGBA", (width, height), (255, 255, 255, 0))
        canvas.alpha_composite(img, ((width - img.width) // 2, (height - img.height) // 2))
        canvas.save(thumb)
    return thumb


def add_flag(ws, row: int, col: int, code: str) -> None:
    if not code:
        return
    flag = FLAG_DIR / f"{code}.png"
    if not flag.exists():
        return
    img = XLImage(str(make_thumbnail(flag)))
    img.anchor = f"{get_column_letter(col)}{row}"
    ws.add_image(img)


def load_rankings() -> dict[str, dict]:
    payload = json.loads(RANKING_PATH.read_text(encoding="utf-8"))
    rankings = {}
    for item in payload["Results"]:
        code = item["IdCountry"]
        rankings[code] = {
            "team": localized_text(item.get("TeamName"), code),
            "rank": item.get("Rank"),
            "points": item.get("DecimalTotalPoints"),
            "confederation": item.get("ConfederationName", ""),
        }
    return rankings


def favorite(code1: str, name1: str, code2: str, name2: str, rankings: dict[str, dict]) -> str:
    rank1 = rankings.get(code1, {}).get("rank")
    rank2 = rankings.get(code2, {}).get("rank")
    if not isinstance(rank1, int) or not isinstance(rank2, int):
        return ""
    if rank1 == rank2:
        return "offen"
    return f"{name1} (RK {rank1})" if rank1 < rank2 else f"{name2} (RK {rank2})"


def uefa_category(code1: str, code2: str) -> str:
    count = int(code1 in UEFA_CODES) + int(code2 in UEFA_CODES)
    if count == 2:
        return "EU team (both)"
    if count == 1:
        return "EU team (one)"
    if code1 and code2:
        return "No EU team"
    return "Offen"


def ends_after_midnight(start: datetime) -> bool:
    clock = start.time()
    return time(0, 0) <= clock < time(8, 0) or clock >= time(22, 15)


def time_bucket(start: datetime) -> str:
    clock = start.time()
    if clock >= time(22, 15):
        return "Spätabend"
    if time(20, 0) <= clock < time(22, 15):
        return "Prime Time"
    if time(17, 0) <= clock < time(20, 0):
        return "Feierabend"
    if time(8, 0) <= clock < time(17, 0):
        return "Tagsüber"
    if time(5, 0) <= clock < time(8, 0):
        return "Morgens"
    return "Nacht"


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1D3557")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def pct(part: int, total: int) -> float:
    return round(part / total * 100, 1) if total else 0.0


def main() -> None:
    matches = sorted(
        json.loads(MATCHES_PATH.read_text(encoding="utf-8"))["Results"],
        key=lambda m: int(m["MatchNumber"]),
    )
    rankings = load_rankings()
    tz = ZoneInfo(TIMEZONE)

    wb = Workbook()
    ws = wb.active
    ws.title = "Smart Spielplan"
    headers = [
        "Spiel",
        "TAG",
        "Datum",
        "Anstoßzeit",
        "Ende ca.",
        "Endet nach Mitternacht",
        "Zeitfenster",
        "Phase",
        "Gruppe",
        "Team1",
        "Flagge 1",
        "RK1",
        "Team2",
        "Flagge 2",
        "RK2",
        "Favorit",
        "EU/UEFA-Kategorie",
        "Stadt",
        "Stadion",
    ]
    ws.append(headers)

    known_rows = []
    for match in matches:
        start = datetime.fromisoformat(match["Date"].replace("Z", "+00:00")).astimezone(tz)
        end = start + timedelta(minutes=MATCH_BLOCK_MINUTES)
        code1 = team_code(match, "Home")
        code2 = team_code(match, "Away")
        name1 = team_name(match, "Home")
        name2 = team_name(match, "Away")
        row = [
            int(match["MatchNumber"]),
            GERMAN_DAYS[start.strftime("%A")],
            start.strftime("%d.%m.%Y"),
            start.strftime("%H:%M"),
            end.strftime("%H:%M"),
            "Ja" if ends_after_midnight(start) else "Nein",
            time_bucket(start),
            stage_name(match),
            group_name(match),
            name1,
            "",
            rankings.get(code1, {}).get("rank", ""),
            name2,
            "",
            rankings.get(code2, {}).get("rank", ""),
            favorite(code1, name1, code2, name2, rankings),
            uefa_category(code1, code2),
            city_name(match),
            stadium_name(match),
        ]
        ws.append(row)
        excel_row = ws.max_row
        ws.row_dimensions[excel_row].height = 24
        add_flag(ws, excel_row, 11, code1)
        add_flag(ws, excel_row, 14, code2)
        if code1 and code2:
            known_rows.append((match, start, code1, name1, code2, name2))

    for col, width in {
        "A": 8, "B": 13, "C": 12, "D": 12, "E": 10, "F": 20, "G": 28,
        "H": 18, "I": 12, "J": 24, "K": 11, "L": 7, "M": 24, "N": 11,
        "O": 7, "P": 24, "Q": 18, "R": 18, "S": 28,
    }.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    style_header(ws)

    fav_ws = wb.create_sheet("Favoritenübersicht")
    fav_ws.append(["Auswertung", "Wert"])
    total = len(matches)
    after = sum(1 for m in matches if ends_after_midnight(datetime.fromisoformat(m["Date"].replace("Z", "+00:00")).astimezone(tz)))
    before = total - after
    fav_ws.append(["Spiele gesamt", total])
    fav_ws.append(["Enden vor Mitternacht", f"{before} ({pct(before, total)}%)"])
    fav_ws.append(["Enden nach Mitternacht", f"{after} ({pct(after, total)}%)"])
    fav_ws.append(["Feststehende Gruppenspiele", len(known_rows)])
    fav_ws.append(["K.-o.-Spiele mit Platzhaltern", total - len(known_rows)])
    fav_ws.append([])
    fav_ws.append(["Team", "Code", "FIFA-Rang", "Ranking-Punkte", "Favorit in Gruppenspielen"])

    fav_counter: Counter[str] = Counter()
    for _, _, code1, name1, code2, name2 in known_rows:
        fav = favorite(code1, name1, code2, name2, rankings)
        if fav:
            fav_counter[fav.split(" (RK ")[0]] += 1
    for code, data in sorted(rankings.items(), key=lambda item: item[1].get("rank") or 999):
        count = fav_counter.get(data["team"], 0)
        if count:
            fav_ws.append([data["team"], code, data["rank"], data["points"], count])
    for col in ["A", "B", "C", "D", "E"]:
        fav_ws.column_dimensions[col].width = 24
    style_header(fav_ws)

    grp_ws = wb.create_sheet("Gruppen-Favoriten")
    grp_ws.append(["Gruppe", "Team", "Code", "FIFA-Rang", "Ranking-Punkte", "Rolle nach Ranking"])
    groups: dict[str, dict[str, str]] = defaultdict(dict)
    for match, _, code1, name1, code2, name2 in known_rows:
        group = group_name(match)
        if group:
            groups[group][code1] = name1
            groups[group][code2] = name2
    for group in sorted(groups):
        ranked = sorted(groups[group].items(), key=lambda item: rankings.get(item[0], {}).get("rank") or 999)
        for pos, (code, name) in enumerate(ranked, 1):
            role = "Gruppenfavorit" if pos == 1 else "Top-2-Kandidat" if pos == 2 else "Außenseiter"
            grp_ws.append([group, name, code, rankings.get(code, {}).get("rank", ""), rankings.get(code, {}).get("points", ""), role])
    for col in ["A", "B", "C", "D", "E", "F"]:
        grp_ws.column_dimensions[col].width = 22
    style_header(grp_ws)

    time_ws = wb.create_sheet("Zeitfenster")
    time_ws.append(["Kategorie", "Spiele", "Vor Mitternacht", "% vor", "Nach Mitternacht", "% nach"])
    buckets: dict[str, Counter] = defaultdict(Counter)
    for match, start, code1, _, code2, _ in known_rows:
        cat = uefa_category(code1, code2)
        key = "Nach" if ends_after_midnight(start) else "Vor"
        buckets[cat][key] += 1
        buckets[cat]["Total"] += 1
    for cat in ["EU team (both)", "EU team (one)", "No EU team"]:
        total_cat = buckets[cat]["Total"]
        before_cat = buckets[cat]["Vor"]
        after_cat = buckets[cat]["Nach"]
        time_ws.append([cat, total_cat, before_cat, pct(before_cat, total_cat), after_cat, pct(after_cat, total_cat)])
    time_ws.append([])
    time_ws.append(["Hinweis", "EU team meint hier UEFA/europäisches Team; K.-o.-Platzhalter sind nicht enthalten."])
    for col in ["A", "B", "C", "D", "E", "F"]:
        time_ws.column_dimensions[col].width = 24
    style_header(time_ws)

    meta = wb.create_sheet("Quelle_Methodik")
    meta.append(["Baustein", "Beschreibung"])
    meta.append(["Spielplan", "FIFA API Cache: wm_2026_matches_fifa.json"])
    meta.append(["Ranking", "FIFA/Coca-Cola Men's World Ranking Cache: fifa_mens_ranking_latest.json"])
    meta.append(["Zeitzone", TIMEZONE])
    meta.append(["Ende ca.", f"Anstoß + {MATCH_BLOCK_MINUTES} Minuten"])
    meta.append(["Zeitfenster", "Nacht 00:00-04:59, Morgens 05:00-07:59, Tagsüber 08:00-16:59, Feierabend 17:00-19:59, Prime Time 20:00-22:14, Spätabend 22:15-23:59."])
    meta.append(["Endet nach Mitternacht", "Ja, wenn Anstoß zwischen 00:00-07:59 liegt oder ab 22:15 startet."])
    meta.append(["EU team", "Sportliche UEFA/europäische Teams, nicht politische EU-Mitgliedschaft."])
    meta.append(["Favorit", "Team mit niedrigerem FIFA-Rang; keine Wettquote und keine echte Prognose."])
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 110
    style_header(meta)

    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
