#!/usr/bin/env python3
"""Build an XLSX schedule for the FIFA World Cup 2026.

The script fetches the official FIFA match calendar and FIFA flag PNGs via
curl, then creates an Excel workbook with embedded flag images.
"""

from __future__ import annotations

import argparse
import json
import subprocess
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


PROJECT_DIR = Path(__file__).resolve().parent
FIFA_MATCHES_URL = (
    "https://api.fifa.com/api/v3/calendar/matches"
    "?language=en&count=500&idCompetition=17&idSeason=285023"
)
FIFA_RANKING_URL = "https://api.fifa.com/api/v3/rankings?gender=1&count=500&language=en"
FLAG_URL = "https://api.fifa.com/api/v3/picture/flags-sq-2/{code}"

GERMAN_DAYS = {
    "Monday": "Montag",
    "Tuesday": "Dienstag",
    "Wednesday": "Mittwoch",
    "Thursday": "Donnerstag",
    "Friday": "Freitag",
    "Saturday": "Samstag",
    "Sunday": "Sonntag",
}


def run_curl(url: str, target: Path | None = None) -> bytes:
    cmd = ["curl", "-sS", "-L", url]
    if target:
        cmd.extend(["-o", str(target)])
        subprocess.run(cmd, check=True)
        return target.read_bytes()
    return subprocess.check_output(cmd)


def localized_text(items: list[dict], default: str = "") -> str:
    for item in items or []:
        if str(item.get("Locale", "")).lower().startswith("en"):
            return item.get("Description", default)
    return items[0].get("Description", default) if items else default


def team_name(match: dict, side: str) -> str:
    team = match.get(side)
    if team:
        return localized_text(team.get("TeamName"), team.get("ShortClubName", ""))
    placeholder = match.get("PlaceHolderA" if side == "Home" else "PlaceHolderB")
    return placeholder or ""


def team_code(match: dict, side: str) -> str:
    team = match.get(side)
    return team.get("IdCountry") if team else ""


def ensure_flag(code: str, flag_dir: Path) -> Path | None:
    if not code:
        return None
    path = flag_dir / f"{code}.png"
    if path.exists() and path.stat().st_size > 0:
        return path
    run_curl(FLAG_URL.format(code=code), path)
    try:
        with PILImage.open(path) as img:
            img.verify()
    except Exception:
        path.unlink(missing_ok=True)
        return None
    return path


def make_thumbnail(src: Path, thumb_dir: Path, width: int = 34, height: int = 24) -> Path:
    thumb = thumb_dir / src.name
    if thumb.exists() and thumb.stat().st_size > 0:
        return thumb
    with PILImage.open(src) as img:
        img = img.convert("RGBA")
        img.thumbnail((width, height), PILImage.LANCZOS)
        canvas = PILImage.new("RGBA", (width, height), (255, 255, 255, 0))
        x = (width - img.width) // 2
        y = (height - img.height) // 2
        canvas.alpha_composite(img, (x, y))
        canvas.save(thumb)
    return thumb


def fetch_matches(cache_path: Path, refresh: bool) -> list[dict]:
    if refresh or not cache_path.exists():
        raw = run_curl(FIFA_MATCHES_URL)
        cache_path.write_bytes(raw)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    matches = payload["Results"]
    return sorted(matches, key=lambda m: int(m["MatchNumber"]))


def fetch_rankings(cache_path: Path, refresh: bool) -> dict[str, dict]:
    if refresh or not cache_path.exists():
        raw = run_curl(FIFA_RANKING_URL)
        cache_path.write_bytes(raw)
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    rankings = {}
    for item in payload.get("Results", []):
        code = item.get("IdCountry")
        if code:
            rankings[code] = {
                "rank": item.get("Rank"),
                "name": localized_text(item.get("TeamName"), code),
                "points": item.get("DecimalTotalPoints"),
            }
    return rankings


def favorite(match: dict, rankings: dict[str, dict]) -> str:
    code1 = team_code(match, "Home")
    code2 = team_code(match, "Away")
    rank1 = rankings.get(code1, {}).get("rank")
    rank2 = rankings.get(code2, {}).get("rank")
    if not isinstance(rank1, int) or not isinstance(rank2, int):
        return ""
    if rank1 == rank2:
        return "offen"
    fav_side = "Home" if rank1 < rank2 else "Away"
    fav_rank = min(rank1, rank2)
    return f"{team_name(match, fav_side)} (RK {fav_rank})"


def build_workbook(
    matches: list[dict],
    rankings: dict[str, dict],
    output_path: Path,
    timezone: str,
) -> None:
    tz = ZoneInfo(timezone)
    flag_dir = PROJECT_DIR / "flags"
    thumb_dir = flag_dir / "_thumbs"
    flag_dir.mkdir(exist_ok=True)
    thumb_dir.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "WM 2026 Spielplan"

    headers = [
        "TAG",
        "Datum",
        "Anstoßzeit",
        "Team1",
        "Flagge Team1",
        "Team2",
        "Flagge Team2",
        "Favorit",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1D3557")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for match in matches:
        kickoff = datetime.fromisoformat(match["Date"].replace("Z", "+00:00")).astimezone(tz)
        team1 = team_name(match, "Home")
        team2 = team_name(match, "Away")
        code1 = team_code(match, "Home")
        code2 = team_code(match, "Away")

        ws.append(
            [
                GERMAN_DAYS[kickoff.strftime("%A")],
                kickoff.strftime("%d.%m.%Y"),
                kickoff.strftime("%H:%M"),
                team1,
                "",
                team2,
                "",
                favorite(match, rankings),
            ]
        )
        ws.row_dimensions[row].height = 24

        for code, col in ((code1, 5), (code2, 7)):
            flag = ensure_flag(code, flag_dir)
            if not flag:
                continue
            thumb = make_thumbnail(flag, thumb_dir)
            img = XLImage(str(thumb))
            img.anchor = f"{get_column_letter(col)}{row}"
            ws.add_image(img)

        row += 1

    widths = {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 24,
        "E": 14,
        "F": 24,
        "G": 14,
        "H": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in cells:
            cell.alignment = Alignment(vertical="center")
        cells[4].alignment = Alignment(horizontal="center", vertical="center")
        cells[6].alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    meta = wb.create_sheet("Quelle")
    meta.append(["Quelle", FIFA_MATCHES_URL])
    meta.append(["Ranking-Quelle", FIFA_RANKING_URL])
    meta.append(["Zeitzone für Anstoßzeit", timezone])
    meta.append(["Spiele", len(matches)])
    meta.append(["Favorit-Logik", "Niedrigerer Rang in der aktuellen FIFA/Coca-Cola-Men's-World-Ranking-API."])
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 120
    for cell in meta[1]:
        cell.font = Font(bold=True)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="WM_2026_Spielplan.xlsx")
    parser.add_argument("--timezone", default="Europe/Berlin")
    parser.add_argument("--refresh", action="store_true", help="Fetch the FIFA API again")
    args = parser.parse_args()

    cache_path = PROJECT_DIR / "wm_2026_matches_fifa.json"
    ranking_cache_path = PROJECT_DIR / "fifa_mens_ranking_latest.json"
    output_path = PROJECT_DIR / args.output
    matches = fetch_matches(cache_path, args.refresh)
    rankings = fetch_rankings(ranking_cache_path, args.refresh)
    build_workbook(matches, rankings, output_path, args.timezone)
    print(f"Wrote {output_path} with {len(matches)} matches.")


if __name__ == "__main__":
    main()
